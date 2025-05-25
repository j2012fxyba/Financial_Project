
import unittest
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from parameterized import parameterized
from selenium.webdriver.common.action_chains import ActionChains 
from openpyxl import load_workbook

#selenium 传入参数化数据
class BMI_test1(unittest.TestCase):


    def load_data(file_path,sheet_name):
        workbook=load_workbook(file_path)
        sheet1=workbook[sheet_name]
        for row in sheet1.iter_rows():
            for cell in row:
                print(cell.value)


    def calcu_bmi(self,weight,height):
        acture_bmi=float(weight) / (height **2)
        a_bmi=round(acture_bmi,1)
        print(a_bmi)
        return a_bmi

    def calcu_catetory(self,a_bmi):
        if 0<a_bmi <18.5:
            return 'underweight'
        elif 18.5<a_bmi<25:
            return 'Normal'
        elif 25<a_bmi<30:
            return 'overweight'
        else:
            return  'obesity'
    
    parameterized.expand(load_data())
    def test_calcu_bmi(self,weight,height,expected_bmi,expected_catetory):
        pass
        
        
        
    
  
    
    def apid_BMI(self):
        driver=webdriver.Chrome()
        url='http://127.0.0.1:5000/'
        driver.get(url)
        print(driver.current_url)
        time.sleep(2)
        print(driver.title)

        a=driver.find_element(By.XPATH,'//div[@class="calculator"]')
        #print(a.get_attribute('outerHTML'))
        z=a.find_element(By.XPATH,'//input[@id="height"]')
        ActionChains(driver).move_to_element(z).click().perform()
        time.sleep(2)
        z.send_keys(self.weight)
        
        w= a.find_element(By.XPATH,'//input[@id="weight"]')
        time.sleep(2)
        w.send_keys(self.height)

        button=driver.find_element_by_tag_name('button')
        time.sleep(1)
        button.click()
        time.sleep(3)
        res=driver.find_element(By.XPATH,'//div[@id="result"]')
        bmi=res.find_element(By.XPATH,'./h3/text')
        print(bmi)
        
        

        driver.quit()
        driver.close()

T=BMI_test1()
T.test_BMI()