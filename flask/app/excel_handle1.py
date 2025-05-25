


from openpyxl import Workbook
from openpyxl import load_workbook

# 创建新的工作簿和工作表 相当于新建excel,此时内存中将生成一个空的excel文件
wb = Workbook()
#wb.create_sheet()  创建新的 sheet 
#wb.active 是获取当前工作簿中 活动的（active）工作表（即打开 Excel 时默认选中的那个 Sheet）
#如果工作簿有多个 Sheet，wb.active 会返回 最后一次操作的工作表（相当于 Excel 中当前显示的 Sheet）
sheet = wb.active

#load_workbook('xxx.xlsx') 加载excel 工作表 前提是已经新建 保存好的。

# 设置工作表标题 即sheet的名称
sheet.title = "Data Sheet"

# 写入数据
sheet['A1'] = "Hello"
sheet.cell(row=1, column=2, value="World")
sheet.cell(row=1,column=3,value='welcome you')


#追加数据
sheet.append(['Name', 'Age', 'Score'])
sheet.append(['Alice', 24, 95])

# 批量添加数据
data = [
    ["Name", "Age","Score"],
    ["Alice", 25,88],
    ["Bob", 30,78]
]
for row in data:
    sheet.append(row)

from openpyxl.styles import Alignment
# 设置A1 单元格居中
sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")


print('得到全部数据')
for row in sheet.iter_rows():
    for cell in row:
        print(cell.value)
        #设置整个单元格内容 都居中，如果数据量很大，可能会影响性能
        cell.alignment = Alignment(horizontal="center", vertical="center")


#索引规则：行号和列号均从 1 开始（与 Excel 一致） 列号可用数字（如 1）或字母（如 "A"）表示

#先插入空行，然后在插入数据
sheet.insert_rows(7)
sheet.cell(row=7,column=1,value='lilih')
sheet.cell(row=7,column=2,value=45)
sheet.cell(row=7,column=3,value=65)

# 方法2：直接操作单元格（无需插入空行）
sheet["A8"] = "uop"
sheet["B8"] = 34
sheet["C8"] = 90

from openpyxl.utils import column_index_from_string
print('修改单元格数据')

#sheet.cell(row=1,column='G',value='New Value') # int and str 格式不支持 instances 

#像 G 列 代码将G与 数字列号 对比，找不到，建议将 G转化为数字
column_num = column_index_from_string('G')  # 返回7
sheet.cell(row=10,column=column_num,value='New Year')


print('第3列数据',sheet['C3'].value)
#cell_2_F=sheet.cell(2,'F')   #这里的F 也不识别，同样需要转化
clou_num=column_index_from_string('F')
cell_2_F=sheet.cell(2,column_num)
print(cell_2_F.value)


#通过循环遍历获取单元格内容
for row1 in range(1,3):
    for col in range(1,4):
        cell=sheet.cell(row=row1,column=col) 
        print(cell.value)   



# 保存时指定文件名和路径
#save_path = "C:\\Users\\YourName\\Documents\\my_excel.xlsx"  # 绝对路径
# 或
save_path = "handel_excel.xlsx"  # 相对路径（相对于当前工作目录）
# 保存文件
wb.save(save_path)  #文件实际在此刻创建 也就是从内存空间的磁盘的文件空间

