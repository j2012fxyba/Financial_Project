# -*- coding: utf-8 -*-

import unittest
from  openpyxl import load_workbook
from parameterized import parameterized
import HtmlTestRunner
from selenium import webdriver


'''
#cls 是类方法 类似于self（@classmethod）中的一个约定命名参数，代表类本身（class itself）
#self 指向实例对象，实例对象是 单个 实例的  而 cls指向类方法 ，类方法是和类绑定的
'''


#从excel 中读取数据 参数化 然后 selenium传入到抓取的元素中
class test_BMI(unittest.TestCase):
    
    @classmethod
    
    def setUpclass(cls):

        print('测试类开始执行')
        #在类方法内部，不能直接访问实例属性（因为还没有实例化） 将dirver与类绑定
        #所以不能是 self.driver
        
        cls.driver=webdriver.Chrome()
        #driver1=webdriver.Chrome()   #区别， driver1只在此方法中有效，而在类别的方法无法访问
        cls.driver.implicitly_wait(10)  #隐士等待 10s
    

    path='D:\\tool\\PythonTest\\flask\\flask_templates\\bmi_data.xlsx'
    sheet_name='bmi_data'
    def load_test_data(path,sheet_name):
        
        row_data=[]
        try:

            workbook=load_workbook(path)   #加载excel表格
            sheet_name1=workbook[sheet_name] #加载excel的sheet_name
            #循环遍历行，从第二行开始 ，跳过标题
            for row in sheet_name1.iter_rows(min_row=2,values_only=True):
                #print(tuple(row)) # 输出: (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, ...)
                #print(len(row))
                
                #para 参数化只接受元组的 数据结构
                process_data=(
                    row[0],
                    float(row[1]),
                    float(row[2]),
                    float(row[3]),
                    row[4])
                
                row_data.append(process_data)
            workbook.close()
            return row_data
            
        except Exception as e:
            print(e)


    def calacuator_bmi(self,height,weight):
        #注意 这里 parametered 会自动将参数传递给方法，不需要其他的引用方式
        #公式
        height_m = float(height)  #搞错了 单元格里面是1.75 不用 /100
        weight1=float(weight)
        acture_bmi = weight1 / (height_m ** 2)
        acture_bmi=round(acture_bmi,2)
    
        return acture_bmi
    
    def get_bmi_category(self, bmi_val):
        """根据BMI值返回分类（固定逻辑）"""
        if bmi_val < 18.5:
            return "underweight"  # 这里返回固定字符串
        elif 18.5 <= bmi_val < 25:
            return "Normal"
        elif 25 <= bmi_val < 30:
            return "overweight"
        else:
            return "obesity"


    @parameterized.expand(load_test_data(path,sheet_name))
    def test_calculate1(self,name, height, weight, expected_bmi, expected_category):

        acture_bmi=self.calacuator_bmi(height,weight)

        self.assertEqual(expected_bmi,acture_bmi,f"计算错误：{expected_bmi}!={acture_bmi}")

        actual_category=self.get_bmi_category(acture_bmi)
        self.assertEqual(expected_category,actual_category,f"分类失败：{expected_category}！={actual_category}")


def main():
        
        #创建一个测试加载器（Test Loader）实例，用于自动发现和加载测试用例。
        #从指定的测试类（TestBMICalculator）中提取所有测试方法，并组装成一个测试套件（TestSuite 对象）。
        #如果不适用测试套件，需要手动添加  
        #suite=unittest.TestSuite() 
        #suite.addTest(test_BMI('test_calculate1'))  #按照 类中的方法手动添加
    '''
        

    自动加载的原理：
    加载器会扫描 test_BMI 类中所有以 test_ 开头的方法（例如 test_bmi_calculation）
    如果使用了 @parameterized.expand，还会为每组参数生成独立的子测试用例。

    '''
        
    suite_1 = unittest.TestLoader().loadTestsFromTestCase(test_BMI)
    

    runner=unittest.TextTestRunner(verbosity=2)
    #result=runner.run(suite_1)

    # #可选择
    #print(result.failfast,result.errors)

    # 运行测试并生成 HTML 报告
  
    runner = HtmlTestRunner.HTMLTestRunner(
    output='reports',  # 报告输出目录
    report_title='测试报告标题',
    add_timestamp=True,

    
    )
    runner.run(suite_1)

if __name__=='__main__':
    main()
