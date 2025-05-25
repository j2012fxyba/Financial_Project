
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

#excel的行和列的 读取方式

class excel_handle():

    def __init__(self,path_file,sheet_name):

        self.workbook=load_workbook(path_file)
        self.sheetname1=self.workbook[sheet_name]
        print(self.sheetname1.max_row)   #打印工作表的行数
        print(self.sheetname1.max_column)  #获取总列数
        #return path_file,sheet_name  __init__是初始化方法，不需要返回return

    def get_data(self):
        print(' #直接遍历 row的 第1 ~第6 范围内数据 前5行')
        for row in range(1,6):  
            for col in range(1,6): 
                cell=self.sheetname1.cell(row=row,column=col)
                print(cell.value)
                

    def get_row_all(self):
        ##values_only=True: 表示只返回单元格的值，而不是单元格对象
        for row in self.sheetname1.iter_rows():
            for cell in row:
                print(cell.value)
        print('直接根据行号和列号获取数据')
        cell=self.sheetname1.cell(2,5)
        print(cell.value)
            
 
        # for hang_1 in sheetname1[1]: #直接打印出第一行  表头从1开始

        #     print (hang_1.value)


                
    #根据行号获取整行数据            
    def row_data(self,row_index):
        row_data=[]
        print('获取指定行 号的数据')
        for row in self.sheetname1[row_index]:
            row_data.append(row.value)
        print(row_data)
        return row_data
        

   
    def get_excel_clo_index(self,row_index,column_index):
        print(' #根据行号+列号 获取数据')
        cell_value=self.sheetname1.cell(row=row_index,column=column_index).value
        print(f"{row_index}行,{column_index}列的值是：{cell_value}")
        return cell_value

    
    def get_excel_row(self,row_index,column_name):
        print('#根据列名获取 列值')
        cell_reference = f"{column_name.upper()}{row_index}"
        cell_value = self.sheetname1[cell_reference].value
        print(f"{row_index}行,{column_name}列单元格值是：{cell_value}")

            
        #方法二
        #使用 openpyxl.utils.column_index_from_string('L') 将字母转换为数字（L→12）
   
        column_index = column_index_from_string(column_name.upper()) 
        cell=self.sheetname1.cell(row=row_index, column=column_index).value
        print(cell)


    def get_cell_value_by_row_and_column(self, row_number, column_identifier):
        """
        根据行号和列标识(名称或索引)获取单元格值
        
        参数:
            file_path: Excel文件路径
            sheet_name: 工作表名称
            row_number: 行号(从1开始)
            column_identifier: 列标识(可以是字母如'A'或数字如1)
        
        返回:
            单元格的值
        """
        
        if isinstance(column_identifier, str):
            # 如果是列名(如'A', 'B')
            column_letter = column_identifier.upper()
            cell_reference = f"{column_letter}{row_number}"
            return self.sheetname1[cell_reference].value
        elif isinstance(column_identifier, int):
            # 如果是列索引(从1开始)
            return self.sheetname1.cell(row=row_number, column=column_identifier).value
        else:
            raise ValueError("column_identifier必须是字符串(列名)或整数(列索引)")




path_file='D:\\tool\\PythonTest\\AutoTest\\seleniums\\hushen300.xlsx'
sheet_name='Sheet1'
T=excel_handle(path_file,sheet_name)  #类中有初始化方法 __init__ 所以实例化时，直接传参

T.get_data()
#T.get_row_all()  #获取所有行
T.row_data(4)   #index =4  因为有表头，所以显示的是第3行数据
#T.get_excel_row(row_index=5,column_name='最低')   #column_name不支持 ’列标题搜索‘
T.get_excel_row(row_index=6,column_name='L')

T.get_excel_clo_index(row_index=7,column_index=8)  #column_index

# 使用示例
value = T.get_cell_value_by_row_and_column(row_number=3,column_identifier='I')  # 或者使用 2
print(value)











    