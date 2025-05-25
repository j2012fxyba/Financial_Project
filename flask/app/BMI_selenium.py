import numpy as np
import unittest
from selenium import webdriver
from selenium.webdriver.common.by import By
import time
from selenium.webdriver.common.action_chains import ActionChains 
from openpyxl import load_workbook
from parameterized import parameterized 
import HTMLTestRunner

#selenium 传入参数化数据
class BMI_test1(unittest.TestCase):

    path='D:\\tool\\PythonTest\\flask\\flask_templates\\bmi_data.xlsx'
    sheet_name='bmi_data'    
    def load_data(path,sheet_name):
        workbook=load_workbook(path)
        sheet1=workbook[sheet_name]
        for row in sheet1.iter_rows():
            for cell in row:
                print(cell.value)

    @parameterized.expand(load_data(path,sheet_name))
    def send_BMI(self,weight,height):
        driver=webdriver.Chrome()
        url='http://127.0.0.1:5000/'
        driver.get(url)
        print(driver.current_url)
        time.sleep(2)
        print(driver.title)

        a=driver.find_element(By.XPATH,'//div[@class="calculator"]')
        #print(a.get_attribute('outerHTML'))
        z=a.find_element(By.XPATH,'//input[@id="height"]')
        ActionChains(driver).move_to_element(z).click().perform()
        time.sleep(2)
        z.send_keys(height)
        
        w= a.find_element(By.XPATH,'//input[@id="weight"]')
        time.sleep(2)
        w.send_keys(weight)

        button=driver.find_element_by_tag_name('button')
        time.sleep(1)
        button.click()
        time.sleep(3)
        res=driver.find_element(By.XPATH,'//div[@id="result"]')
        self.bmi=res.find_element(By.XPATH,'./h3/text')
        #self.fenlei=res.find_element(By.XPATH,'./h3/text')
      
        return self.bmi
    
    def get_bmi(self,weight,height):
        acture_bmi=float(weight) / (height **2)
        a_bmi=round(acture_bmi,1)
        print(a_bmi)
        return a_bmi
    
    def get_catetory(a_bmi):
        if 0<a_bmi <18.5:
            return 'underweight'
        elif 18.5<a_bmi<25:
            return 'Normal'
        elif 25<a_bmi<30:
            return 'overweight'
        else:
            return  'obesity'
        
    def test_calcu(self,expected_bmi,expected_catetory):
        self.assertEqual(expected_bmi,self.bmi,f"计算错误：{expected_bmi}!={self.bmi}")

        actual_category=self.get_catetory(self.bmi)
        self.assertEqual(expected_catetory,actual_category,f"分类失败：{expected_catetory}！={actual_category}")


def main():
    suite1=unittest.TestLoader().loadTestsFromModule(BMI_test1)

    runner = HTMLTestRunner.HTMLTestRunner(
    output='reports',  # 报告输出目录
    report_title='测试报告标题',
    add_timestamp=True,
    )
    runner.run(suite1)    

T=BMI_test1()

T.load_data()
T.send_BMI()